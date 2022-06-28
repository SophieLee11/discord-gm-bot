import discord
import datetime
from openpyxl import load_workbook



async def get_message_list(client, guild_id, days, phrase):
	guild = client.get_guild(int(guild_id))
	guild_members_count = guild.member_count
	channels = [i.id for i in guild.text_channels]

	limit = 10000
	message_list = []
	yesterday = datetime.datetime.today() - datetime.timedelta(days = days)

	for channel_id in channels:
		try:
			channel = client.get_channel(int(channel_id))
			messages = channel.history(limit = limit, after = yesterday)

			async for message in messages:
				if phrase == message.content:
					message_list.append(message.author.id)
						
		except Exception as e:
		 	pass

	message_list = set(message_list)

	return guild, message_list, guild_members_count





def get_list_from_txt():
	server_list = []
	servers = open('servers.txt', 'r').read().split('\n')
	for server in servers:
		if len(server) > 0:
			is_server_ok = True
			for symbol in server:
				if not symbol.isdigit():
					is_server_ok = False
			if is_server_ok:
				server_list.append(server)
			else:
				print('Incorrect server ID in servers.txt')

	return server_list




def get_checklist(client, server_list):
	text = 'Checklist 📃:\n\n'
	names = []

	longest_name = 0
	for guild_id in server_list:
		guild = client.get_guild(int(guild_id))
		names.append(guild.name)
		if len(guild.name) > longest_name:
			longest_name = len(guild.name)

	for i in range(len(names)):
		name = names[i] + (longest_name - len(names[i])) * ' '
		guild_id = server_list[i]
		text += f"{i + 1}. {name} [{guild_id}]\n"

	return text




def save_txt_from_list(server_list):
	servers = open('servers.txt', 'w')

	for server in server_list:
		servers.write(server + '\n')




def add_to_xlsx_file(server_name, server_id, value, gm_amount, members_amount):

	wb = load_workbook(filename = "data.xlsx")
	ws = wb['Data']

	table = ws.tables["All"]

	new_row = int(table.ref[table.ref.index('F') + 1:]) + 1

	date = datetime.date.today()
	
	ws[f'A{new_row}'] = date
	ws[f'B{new_row}'] = server_name
	ws[f'C{new_row}'] = server_id
	ws[f'D{new_row}'] = value
	ws[f'E{new_row}'] = gm_amount
	ws[f'F{new_row}'] = members_amount

	table.ref = f'A1:F{new_row}'

	wb.save('data.xlsx')



def get_cur_dt(full = False):
	if full:
		return datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
	else:
		return datetime.datetime.now().strftime('%d.%m.%Y')